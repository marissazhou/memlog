"""File relevant operations class
   This class includes all file relevant operations 

.. module:: common 
   :platform: Ubuntu Unix
   :synopsis: A module for all type file relevant process 

.. moduleauthor:: Lijuan Marissa Zhou <marissa.zhou.cn@gmail.com>

"""
import datetime
from scipy.stats import norm,lognorm
import math
import scipy
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

from fileuploader.models import Sensor
from fileuploader.models import Picture
from fileuploader.models import Album
from fileuploader.models import SensorFile

from memlog_settings import * 
from utils import *
from db.fileuploaderHelper import *

"""
 	constants for distritubion types	
"""
DIST_TYPE_NORMAL 	= 0

def walk_sensor_file(sensorfile, uid, device_type, image_ids):
	""" This function walks through sensor file and store all sensor information into database 
	For autographer sensor file:
		1. ignore the first two lines

	:param  sensorfile: full path of sensor file that is posted and stored in the server 
	:type   sensorfile: string 
	:param  uid: user id 
	:type   uid: long 
	:param  device_type: device type check with settings.py file
	:type   device_type: int 
	:param  image_ids: image id map dictionary 
	:type   image_ids: dict 
	"""
	if device_type is DEVICE_TYPE_AUTOGRAPHER:
		process_autographer_sensor_file(sensorfile, uid, image_ids)
		print 'DEVICE_TYPE_Autographer'
	elif device_type is DEVICE_TYPE_SENSECAM:
		print 'DEVICE_TYPE_sensecam'
	if device_type is DEVICE_TYPE_ANDROID_PHONE:
		print 'DEVICE_TYPE_ANDROID_PHONE'

def process_autographer_sensor_file(filepath, uid, image_ids):
	""" This function process autographer sensor file 

	:param  filepath: 
	:type   filepath: string 
	"""
	row_idx			= 1
	col_idx			= 1
	"""
	wb 				= load_workbook(filename = sensorfile)
	sheet_ranges 	= wb['sheet1']
	# read row
	print type(sheet_ranges)
	while True:
		if sheet_ranges['A1'] is not None:
			# read column
			while True:
				col 		= get_column_letter(col_idx)
				cell_value 	= sheet_ranges[col+'1'].value
				sensor_value= []
				capture_time= datetime.now()
				image_name 	= ''
				if cell_value is not None:
					print cell_value 
					if col_idx is 1:
						capture_time = autographer_string_to_datetime(cell_value) 
					elif col_idx is 2:
						file_name = cell_value
					else: # index>=3
						insert_into_sensor_db(col_idx, cell_value, device_type, image_ids, image_name, capture_time)
					col_idx += 1
				else:
					break
			row_idx += 1
			col_idx	= 1
		else:
			break
	"""
	with open(filepath, 'r') as sensor_file: 
		sensor_recordings = sensor_file.readlines()
		count = 0 
		for line in sensor_recordings:
			row_idx	+= 1
			if row_idx < 4:
				continue
			else:
				print 'Line'
				print line
				one_record = line.split(',')
				capture_time 	= autographer_string_to_datetime(one_record[0])
				image_name		= one_record[1]
				col_idx = 2 
				for cell_value in one_record[2:]:
					cell_value 	= cell_value.strip() # trim string
					col_idx 	+= 1
					insert_into_sensor_db(uid, col_idx, cell_value, DEVICE_TYPE_AUTOGRAPHER, image_ids, image_name, capture_time)
				print one_record
		return

	print 'Matrix of the sensor file:' 
	print row_inx
	print col_inx

def insert_into_sensor_db(uid, index, value, device_type, image_ids, image, capture_time):
	""" This function insert the current record of sensor into sensor table according to sensor type and device type 
	For autographer sensor file:
		1. each column is one sensor values 

	:param  index: index in sensor file, in Autographer sensor file, it is the column of sensor file; in sensecam sensor file, it is...; in android phone, it is... 
	:type   index: int 
	:param  value: sensor value
	:type   value: string 
	:param  device_type: device type check with settings.py file
	:type   device_type: int 
	"""
	image = image+"4.JPG"
	img_id			= -1 # the image was not captured or was deleted 
	img_obj			= None
	if image in image_ids:
		img_id			= image_ids[image]
		img_obj			=Picture(id=img_id)
	sensor_type = get_sensor_type(device_type, index)
	# autographer_string_to_time(time_str)
	if device_type is DEVICE_TYPE_AUTOGRAPHER:
		new_file    = Sensor(user=User(id=uid),img=img_obj,type=sensor_type, \
							capture_at=capture_time)
		new_file.save()

def get_sensor_type(device_type, index):
	""" This function insert the current record of sensor into sensor table according to sensor type and device type 
	For autographer sensor file:
		1. each column is one sensor values 

	:param  index: index in sensor file, in Autographer sensor file, it is the column of sensor file; in sensecam sensor file, it is...; in android phone, it is... 
	:type   index: int 
	:param  device_type: device type check with settings.py file
	:type   device_type: int 

	:return: sensor type id in sensor table. if this sensor is not find, then add into this table
	"""
	sensor_type = -1
	if device_type is DEVICE_TYPE_AUTOGRAPHER:
		abbr = SENSOR_ABBREVIATION_AUTOGRAPHER[str(index)]
		sensor_type = fileuploader_get_sensor_type(abbr, None)	
	return sensor_type

